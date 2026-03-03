"""
식약처 e약은요 API + Excel 병행 — 의약품 지식 JSON 생성 스크립트.

사용법:
    python scripts/fetch_drug_info.py                 # 전체 수집
    python scripts/fetch_drug_info.py --limit 50      # 50건만 테스트
    python scripts/fetch_drug_info.py --resume         # 체크포인트 이어서
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

# ── 경로 설정 ────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = (
    ROOT
    / "경구약제 이미지 데이터(데이터 설명서, 경구약제 리스트)"
    / "단일 경구약제_5,000종 리스트.xlsx"
)
OUTPUT_PATH = ROOT / "data" / "knowledge_base.json"
CHECKPOINT_PATH = ROOT / "data" / "_checkpoint.json"

# ── 식약처 API ────────────────────────────────────
API_BASE = "https://apis.data.go.kr/1471000/DrbEasyDrugInfoService/getDrbEasyDrugList"

# 응답 필드 매핑 (API 필드 → 지식 JSON 필드)
FIELD_MAP = {
    "efcyQesitm": "efficacy",
    "useMethodQesitm": "dosage",
    "atpnQesitm": "precautions",
    "atpnWarnQesitm": "warnings",
    "intrcQesitm": "interactions",
    "seQesitm": "side_effects",
    "depositMethodQesitm": "storage",
}

# ── 로깅 ──────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("fetch_drug_info")


# ── 헬퍼 ──────────────────────────────────────────
def load_env() -> dict[str, str]:
    """ROOT/.env 파일에서 키=값을 읽어 dict로 반환."""
    env: dict[str, str] = {}
    env_path = ROOT / ".env"
    if not env_path.exists():
        return env
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def name_to_medication_id(c_code: str, name: str) -> str:
    """C-Code + 제품명 → medication_id (계약서 규격: 영문대문자+언더스코어)."""
    code = c_code.replace("'", "").replace("-", "_").upper()
    clean = re.sub(r"\(.*?\)", "", name).strip()
    clean = re.sub(r"\s+", "_", clean)
    clean = re.sub(r"[^\w]", "", clean)
    return f"{code}_{clean.upper()}"


def strip_html(text: str | None) -> str:
    """HTML 태그 제거 및 공백 정리."""
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _build_search_names(product_name: str) -> list[str]:
    """제품명에서 여러 검색 키워드를 생성 (성공률 극대화)."""
    candidates: list[str] = []
    name = product_name.strip()

    # 1) 괄호(제조사) 제거
    no_paren = re.sub(r"\s*\(.*?\)", "", name).strip()

    # 2) 숫자+단위 꼬리 제거 (예: 200T, 10C, 480MG 등)
    no_tail = re.sub(r"\s*\d+\s*[TCcMmGg]+.*$", "", no_paren).strip()

    # 3) 숫자+mg/정/캡슐 앞까지만 (핵심 약명)
    core = re.sub(r"[\d].*$", "", no_paren).strip()
    # '정', '캡슐', '산' 등 제형 접미사 포함
    if core and core[-1] not in ("정", "산", "액", "셀"):
        core_with_form = core
    else:
        core_with_form = core

    # 중복 제거하면서 순서 유지
    seen: set[str] = set()
    for c in [no_paren, no_tail, core_with_form, name]:
        c = c.strip()
        if c and c not in seen:
            candidates.append(c)
            seen.add(c)

    return candidates


def fetch_drug_detail(api_key: str, product_name: str) -> dict | None:
    """e약은요 API로 약품 상세 정보를 조회. 여러 검색 전략 시도."""
    search_names = _build_search_names(product_name)

    for search_name in search_names:
        try:
            params = {
                "serviceKey": api_key,
                "itemName": search_name,
                "type": "json",
                "numOfRows": "3",
            }
            url = f"{API_BASE}?{urllib.parse.urlencode(params, safe='%+')}"
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))

            items = data.get("body", {}).get("items", [])
            if not items:
                continue

            item = items[0]
            result = {
                "api_item_name": item.get("itemName", ""),
                "company": item.get("entpName", ""),
                "item_seq": item.get("itemSeq", ""),
            }
            for api_field, our_field in FIELD_MAP.items():
                result[our_field] = strip_html(item.get(api_field))
            return result

        except Exception:
            continue

    return None


def read_excel() -> list[dict]:
    """Excel 양쪽 시트에서 약품 목록을 읽어온다 (중복 c_code 제거)."""
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    seen_codes: set[str] = set()
    medications: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            c_code = str(row[0]).replace("'", "").strip() if row[0] else ""
            category = str(row[1]).replace("'", "").strip() if row[1] else ""
            product_name = str(row[2]).replace("'", "").strip() if row[2] else ""

            if not c_code or not product_name or c_code in seen_codes:
                continue

            seen_codes.add(c_code)
            medications.append({
                "c_code": c_code,
                "category": category,
                "name": product_name,
            })

    wb.close()
    log.info("Excel에서 %d건 로드 (시트 %d개)", len(medications), len(wb.sheetnames))
    return medications


def load_checkpoint() -> set[str]:
    if CHECKPOINT_PATH.exists():
        data = json.loads(CHECKPOINT_PATH.read_text(encoding="utf-8"))
        return set(data.get("completed", []))
    return set()


def save_checkpoint(completed: set[str]) -> None:
    CHECKPOINT_PATH.write_text(
        json.dumps({"completed": sorted(completed)}, ensure_ascii=False),
        encoding="utf-8",
    )


def _save_output(knowledge: list[dict]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(
            {"medications": knowledge, "total": len(knowledge)},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="식약처 API 의약품 지식 수집")
    parser.add_argument("--limit", type=int, default=0, help="수집 건수 제한 (0=전체)")
    parser.add_argument("--resume", action="store_true", help="체크포인트에서 이어서")
    parser.add_argument("--sleep", type=float, default=0.3, help="API 호출 간격(초)")
    args = parser.parse_args()

    env = load_env()
    api_key = env.get("DATA_GO_KR_API_KEY_ENCODED", "")
    if not api_key:
        log.error("DATA_GO_KR_API_KEY_ENCODED가 .env에 없습니다")
        sys.exit(1)

    medications = read_excel()

    completed = load_checkpoint() if args.resume else set()
    if completed:
        log.info("체크포인트: %d건 이미 처리됨", len(completed))

    knowledge: list[dict] = []
    if args.resume and OUTPUT_PATH.exists():
        existing = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
        knowledge = existing.get("medications", [])
        log.info("기존 결과 %d건 로드", len(knowledge))

    target = medications[: args.limit] if args.limit else medications
    success_count = 0
    fail_count = 0

    for i, med in enumerate(target):
        if med["c_code"] in completed:
            continue

        detail = fetch_drug_detail(api_key, med["name"])

        entry = {
            "medication_id": name_to_medication_id(med["c_code"], med["name"]),
            "name": med["name"],
            "category": med["category"],
            "c_code": med["c_code"],
            "efficacy": "",
            "dosage": "",
            "precautions": "",
            "warnings": "",
            "interactions": "",
            "side_effects": "",
            "storage": "",
            "source": "식약처 의약품안전나라",
        }

        if detail:
            for field in FIELD_MAP.values():
                if detail.get(field):
                    entry[field] = detail[field]
            entry["api_item_name"] = detail.get("api_item_name", "")
            entry["company"] = detail.get("company", "")
            success_count += 1
        else:
            fail_count += 1

        knowledge.append(entry)
        completed.add(med["c_code"])

        processed = i + 1
        if processed % 50 == 0:
            log.info(
                "진행: %d/%d (성공: %d, 실패: %d)",
                processed, len(target), success_count, fail_count,
            )
            _save_output(knowledge)
            save_checkpoint(completed)

        time.sleep(args.sleep)

    _save_output(knowledge)
    save_checkpoint(completed)

    log.info(
        "완료! 총 %d건 (API 성공: %d, API 없음: %d)",
        len(knowledge), success_count, fail_count,
    )
    log.info("저장 위치: %s", OUTPUT_PATH)


if __name__ == "__main__":
    main()
